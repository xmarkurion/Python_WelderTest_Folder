from __future__ import print_function
from mailmerge import MailMerge
from time import gmtime, strftime
from datetime import datetime, timedelta
from configparser import SafeConfigParser
import os
import shutil
import glob
import datetime
import sys
import openpyxl
import configparser


#----- FUNC BLOCK BEGIN


def logo():
    print(
        '\n'
        '███╗   ███╗ █████╗ ██████╗ ██╗  ██╗██╗   ██╗██████╗ ██╗ ██████╗ ███╗   ██╗ \n'   
        '████╗ ████║██╔══██╗██╔══██╗██║ ██╔╝██║   ██║██╔══██╗██║██╔═══██╗████╗  ██║  \n'  
        '██╔████╔██║███████║██████╔╝█████╔╝ ██║   ██║██████╔╝██║██║   ██║██╔██╗ ██║  \n'   
        '██║╚██╔╝██║██╔══██║██╔══██╗██╔═██╗ ██║   ██║██╔══██╗██║██║   ██║██║╚██╗██║  \n'   
        '██║ ╚═╝ ██║██║  ██║██║  ██║██║  ██╗╚██████╔╝██║  ██║██║╚██████╔╝██║ ╚████║  \n'   
        '╚═╝     ╚═╝╚═╝  ╚═╝╚═╝  ╚═╝╚═╝  ╚═╝ ╚═════╝ ╚═╝  ╚═╝╚═╝ ╚═════╝ ╚═╝  ╚═══╝  \n'   
        ' ----------------Welder Test Folder Structure Editor---------------------   \n'                                                                                                                                                
    )

def make_space():
    os.system('cls')
    logo()


def create_configuration_file(folder, date, client_name, job_no, report_no, test_type, report_file_name):
    date_formated = date[0:2] + " " + date[2:-2] + " " + date[-2:]
    working_folder = folder + "/" + date
    os.chdir(working_folder)
    config_file_creator = open('test_data.ini', 'w+')
    config_file_creator.close()

    config = configparser.ConfigParser()

    config.add_section('TEST')
    with open('test_data.ini', 'w') as configfile:  #Save
        config.write(configfile)

    config.read('welder_test.ini')
    config['TEST']['client'] = client_name
    config['TEST']['date'] = date
    config['TEST']['date_formated'] = date_formated.upper()
    config['TEST']['job_no'] = job_no
    config['TEST']['report_no'] = report_no
    config['TEST']['test_type'] = test_type
    config['TEST']['report_name'] = report_file_name
    config['TEST']['client_adr'] = ''

    with open('test_data.ini', 'w') as configfile:
        config.write(configfile)


def fracture_word_editor(word_file, folder, date, client_name, job_no, report_no):
    resurces_folder_name = sys.path[0] + "/Resources"
    word_reports_folder = folder + "/" + date + "/Reports"

    #Date in format XX XXXX XX
    date_formated = date[0:2] + " " + date[2:-2] + " " + date[-2:]

    os.chdir(resurces_folder_name)
    doc = MailMerge('Fracture.docx')
    doc.merge(
        Job_no = job_no,
        Report_no = report_no,
        client = client_name,
        date = date_formated.upper()
    )

    os.chdir(word_reports_folder)
    doc.write(word_file)   

def macro_word_editor(word_file, folder, date, client_name, job_no, report_no):
    resurces_folder_name = sys.path[0] + "/Resources"
    word_reports_folder = folder + "/" + date + "/Reports"
    date_formated = date[0:2] + " " + date[2:-2] + " " + date[-2:]

    os.chdir(resurces_folder_name)
    doc = MailMerge('Macro.docx')
    doc.merge(
        Job_no = job_no,
        Report_no = report_no,
        client = client_name,
        date = date_formated.upper()
    )
    os.chdir(word_reports_folder)
    doc.write(word_file)  

def excel_rename_req_file(excel_file,folder,date,client_name):
    work_folder = folder + "/" + date
    file_name = "REQ " + date.upper() + " " + client_name + ".xlsx"
    os.chdir(work_folder)
    wb = openpyxl.load_workbook(excel_file)
    wb.save(file_name)
    os.remove('REQ.xlsx')

def excel_cert_data_editor(excel_file, folder, date, client_name, test_type):
    work_folder = folder + "/" + date + "/Certs"

    os.chdir(work_folder)
    wb = openpyxl.load_workbook(excel_file)

    SheetTemplate = wb['Template']
    SheetTemplate['F20'] = client_name
    SheetTemplate['O47'] = strftime("%d/%m/%Y.", gmtime())

    if test_type == '2':
        SheetTemplate['F48'] = "-"
        SheetTemplate['I48'] = "N/A"
        SheetTemplate['F47'] = "Yes"
        SheetTemplate['I47'] = "-"

    wb.save(excel_file)

    SheetWelders = wb['Welders']  
    SheetWelders['S3'] = ''.join([c for c in client_name if c.isupper()])  #Generata from Upercase
    SheetWelders['S2'] = strftime("%m%y", gmtime())
    wb.save(excel_file)
    

def excel_req_data_editor(excel_file, folder, date, client_name, job_no, report_no, test_type):
    
    time_string_gen = date
    work_folder = folder + "/" + time_string_gen

    date_formated = date[0:2] + " " + date[2:-2] + " " + date[-2:]

    os.chdir(work_folder)   #Select work Folder
    wb = openpyxl.load_workbook(excel_file)  #open excel file
    
    sheet = wb['Sheet1'] # open sheet for reading or editing

    if(test_type == 1):
        sheet['D1'].value = "FRACTURE WELDING SURVEYOR'S REPORT"
    if(test_type == 2):
        sheet['D1'].value = "MACRO WELDING SURVEYOR'S REPORT"

    sheet['D2'].value = client_name + " " + date_formated.upper() #making changes
    sheet['J1'].value = job_no
    sheet['M1'].value = report_no
    wb.save(excel_file)
    

def create_folder_structure(type,folder,date):
    
    resurces_folder_name = sys.path[0] + "/Resources"
    excel_req_file_name = resurces_folder_name + "/REQ.xlsx"
    excel_certgen_file_name = resurces_folder_name + "/cert_gen.xlsx"
    # word_fracture_file_name = resurces_folder_name + "/Fracture.docx"

    #Genereate the folder name as exammple 13Jun19
    time_string_gen = date

    #Open the path created by User. 
    os.chdir(folder)

    if not (os.path.isdir(time_string_gen)):
        os.mkdir(time_string_gen)
    
    #Folder Name redirect's
    base_foler_name = folder + "/" + time_string_gen
    inside_base_folder_name = base_foler_name
    inside_cers_folder_name = inside_base_folder_name + "/Certs"

    os.chdir(inside_base_folder_name)
    if not(os.path.isdir("Certs")):
        os.mkdir("Certs")

    if not(os.path.isdir("Reports")):
        os.mkdir("Reports")

    if not(os.path.isdir("Welders")):
        os.mkdir("Welders")

    if(type == 1):
        if not(os.path.isdir("Fracture")):
            os.mkdir("Fracture")

    if(type == 2):
        if not(os.path.isdir("Macro")):
            os.mkdir("Macro")
        
    #Here Script copy the REQ.xls from resources file.
    if not(os.path.isfile("REQ.xlsx")):
        shutil.copyfile(excel_req_file_name,"REQ.xlsx")

    #Here Script copy the Cert_gen.xls from resources file.
    os.chdir(inside_cers_folder_name)
    if not(os.path.isfile("cert_gen.xlsx")):
        shutil.copyfile(excel_certgen_file_name,"cert_gen.xlsx")

    #Here Script copy Reports into reperts acording to choice.
    #os.chdir(inside_reports_folder_name)
    #if not(os.path.isfile("Fracture.docx")
    #    shutil.copyfile(word_fracture_file_name,"Fracture.docx")
   
    

#----- FUNC BLOCK ENDS

# - 1 For Fracture  - 2 For Macro
logo()

folder_name = input('Please enter folder path: ')

make_space()

print('Select test type: \n 1.Fracture \n 2.Macro \n')
test_type = input('Enter choice: ')

make_space()

print('Do you want to use today date for folder & reports: ' + strftime("%d%b%y", gmtime()) + '\n 1. For Yes \n 2. For NO.' )
folder_date = input('\n Enter choice: ')

if(int(folder_date) == 2):
    make_space()
    print('Your own date in format DayMonthYear ')
    folder_custom_date = input('Please enter the date: ')

make_space()

client_name = input('Enter client name: ')

make_space()

print('Please enter numbers: ')
entered_job_no = input('Enter Job Number: ')
entered_report_no = input('Entered Report Number: ')

make_space()

print(' \n The data entered as folow: \n')
print('Folder name: ' + folder_name + ' \n')
print('Test type: ' + test_type + ' \n')
print('Job Number: ' + entered_job_no + '\n')
print('Report Number: ' + entered_report_no + ' \n')
print('Will be saved .... \n \n')

print('Creating folder structure in: ' + folder_name + '\n ---- ^^_^^ -----')

#FRACTURE SECTION
if(int(folder_date) == 1):
    now_date = strftime("%d%b%y", gmtime())
    create_folder_structure(int(test_type),folder_name,now_date)
    print('Editing excel REQ data... \n ---- ^^_^^ -----')
    excel_req_data_editor('REQ.xlsx',folder_name, now_date, client_name,entered_job_no,entered_report_no,int(test_type))
    print('Editing excel Certyficates data... \n ---- ^^_^^ -----')
    excel_cert_data_editor('cert_gen.xlsx',folder_name,now_date,client_name,test_type)
    print('Editing word report data... \n ---- ^^_^^ -----')
    fracture_file_name = "Report " + entered_report_no + " Fracture " + client_name + " Job " + entered_job_no + ".docx"
    fracture_word_editor(fracture_file_name, folder_name, now_date, client_name, entered_job_no,entered_report_no)
    print('Rename REQ File... \n ---- ^^_^^ -----')
    excel_rename_req_file("REQ.xlsx", folder_name, now_date, client_name)
    print('Create and modyfi configuration File... \n ---- ^^_^^ -----')
    create_configuration_file(folder_name, now_date, client_name, entered_job_no, entered_report_no,test_type,fracture_file_name)
    
#MACRO SECTION
if(int(folder_date) == 2):
    create_folder_structure(int(test_type),folder_name,folder_custom_date)
    print('Editing excel REQ data... \n ---- ^^_^^ -----')
    excel_req_data_editor('REQ.xlsx',folder_name, folder_custom_date, client_name,entered_job_no,entered_report_no,int(test_type))
    print('Editing excel Certyficates data... \n ---- ^^_^^ -----')
    excel_cert_data_editor('cert_gen.xlsx',folder_name,folder_custom_date,client_name,test_type)
    print('Editing word report data... \n ---- ^^_^^ -----')
    macro_file_name = "Report " + entered_report_no + " Macro " + client_name + " Job " + entered_job_no + ".docx"
    macro_word_editor(macro_file_name, folder_name, folder_custom_date, client_name,entered_job_no,entered_report_no)
    print('Rename REQ File... \n ---- ^^_^^ -----')
    excel_rename_req_file("REQ.xlsx", folder_name, folder_custom_date, client_name)
    print('Create and modyfi configuration File... \n ---- ^^_^^ -----')
    create_configuration_file(folder_name, folder_custom_date, client_name, entered_job_no, entered_report_no,test_type, macro_file_name)


print(" \n \n ---- DONE -----")
os.system("pause")
